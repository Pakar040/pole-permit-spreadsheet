from typing import List, Dict
import pandas as pd
from model.pole import Pole
import openpyxl
from abc import ABC, abstractmethod


# ----- Static Methods ----- #
def select_template(template: str) -> 'ExcelManager':
    """Creates a manager instance of chosen jurisdiction"""
    if template == 'PSE':
        return PSEManager()


class ExcelManager(ABC):
    """
    Extracts data from excel and formats it

    self.file_path: This should lead to an Excel spreadsheet that uses a template
    self.template_path: This should lead to a template Excel spreadsheet
    self.df: Stores the spreadsheet data in a DataFrame to be used
    self.attachment_map: Dictionary that maps column headers to headers the PoleManager class can understand
    self.reversed_map: Dictionary that maps standard headers that the PoleManager class understands back to template
    headers
    self.template_header: Stores template column headers in a list in the same order they are in the template ot be
    used to format other DataFrames to fit in the template DataFrame
    self.standard_headers: Stores template headers mapped to standard

    Example of use:
    # Read excel data and format for PoleManager
    excel_manager = em.select_jurisdiction('PSE')
    excel_manager.set_file_path('user_input/pse_ground_mold.xlsx')
    excel_manager.read_excel()
    excel_manager.format()
    excel_manager.parse_column('additional_measurements')

    # Extract poles and create make ready
    poles = PoleManager()
    poles.extract_poles(excel_manager.df)
    poles.get_all_violations()

    # Update make ready in spreadsheet and create output
    excel_manager.update_make_ready(poles.pole_list)
    excel_manager.parse_column('additional_measurements')
    excel_manager.format()
    excel_manager.create_output()
    """

    def __init__(self):
        self.file_path = None
        self.template_path = None
        self.df = None
        self.attachment_map: Dict[str] = None
        self.reversed_map: Dict[str] = None
        self.template_headers: List[str] = None
        self.standard_headers: List[str] = None

    def __repr__(self):
        return f"{self.df.to_string(index=False)}"

    def parse_column(self, column: str) -> None:
        """Splits the notes up into attachment dictionaries"""
        for index, row in self.df.iterrows():
            note = row[column]
            notes = note.split('\n')
            parsed_note = []
            for piece in notes:
                part = piece.split(': ')
                try:
                    if part[0] != 'nan':
                        pair = {
                            'name': part[0],
                            'value': part[1],
                        }
                        parsed_note.append(pair)
                except IndexError:
                    parsed_note.append(piece)
            self.df.at[index, column] = parsed_note

    def reverse_parse_column(self, column: str) -> None:
        """Combines attachment dictionaries into a single notes field"""
        for index, row in self.df.iterrows():
            parsed_note = row[column]
            if isinstance(parsed_note, list):
                if len(parsed_note) == 0:
                    self.df.at[index, column] = 'nan'
                else:
                    note = ""
                    for item in parsed_note:
                        if isinstance(item, dict):
                            name = item.get('name', '')
                            value = item.get('value', '')
                            note += f"{name}: {value}\n"
                        elif isinstance(item, str):
                            note += item + "\n"
                    self.df.at[index, column] = note.strip()

    def rename_header(self, attachment_name: str) -> str:
        """Renames attachments to match standard convention or template convention"""
        try:
            # Try to get the value from the original map
            return self.attachment_map[attachment_name]
        except KeyError:
            try:
                # If not found in the original map, try the reversed map
                return self.reversed_map[attachment_name]
            except KeyError:
                # If the key is not found in either map, return the original attachment name
                return attachment_name

    @abstractmethod
    def set_file_path(self, file_path: str) -> None:
        pass

    @abstractmethod
    def update_make_ready(self, lst: List[Pole]) -> None:
        pass

    @abstractmethod
    def read_data_frame(self, data_frame) -> None:
        pass

    @abstractmethod
    def read_excel(self) -> None:
        pass

    @abstractmethod
    def read_excel_in_read_mode(self) -> None:
        pass

    @abstractmethod
    def format(self) -> None:
        pass

    @abstractmethod
    def create_output(self, file_path: str) -> None:
        pass


class PSEManager(ExcelManager):
    """Extracts data from excel and formats it to PSE template"""
    def __init__(self):
        super().__init__()
        self.template_path = 'model/templates/PSE.xlsx'
        self.attachment_map = {
            'Seq #': '_title',
            'PSE Pole #': 'pse_tag_number',
            'Pole Type T/D': 'pse_pole_type',
            'Pole Owner': 'jursidiction',
            'PSE Umap': 'PSE Umap',
            'Location': 'Location',
            'City/Area': 'City/Area',
            'Neutral': 'neutral_height',
            'Secondary': 'secondary_spool.txt',
            'Drip Loop': 'drip_loop',
            'Secondary Riser': 'secondary_riser',
            'Street Light': 'streetlight',
            'CATV': 'catv',
            'TelCo': 'telco',
            'Fiber': 'fiber',
            'Request Attachment': 'Request Attachment',
            'Make Ready Notes': 'make_ready',  # Not in the spreadsheet but is custom
            'Mid Span Violation Notes': 'Mid Span Violation Notes',
            'PSE Field Notes': 'additional_measurements'
        }
        self.reversed_map = {v: k for k, v in self.attachment_map.items()}
        self.template_headers = [
            'Seq #', 'PSE Pole #', 'Pole Type T/D', 'Pole Owner', 'PSE Umap', 'Location',
            'City/Area', 'Neutral', 'Secondary', 'Drip Loop', 'Secondary Riser',
            'Street Light', 'CATV', 'TelCo', 'Fiber', 'Requested Attachment',
            'Make Ready Notes', 'Mid Span Violation Notes', 'PSE Field Notes'
        ]
        self.standard_headers = [self.rename_header(header) for header in self.template_headers]

    def set_file_path(self, file_path: str) -> None:
        """Sets the file path to get DataFrame from"""
        self.file_path = file_path

    def update_make_ready(self, lst: List[Pole]) -> None:
        """Receives a list of poles and adds the make ready data to DataFrame"""
        for index, row in self.df.iterrows():
            if self.df.at[index, 'make_ready'] == 'nan':
                self.df.at[index, 'make_ready'] = lst[index].make_ready
            elif lst[index].make_ready == "":
                pass
            else:
                self.df.at[index, 'make_ready'] += "\n" + lst[index].make_ready

    def read_data_frame(self, data_frame) -> None:
        """Takes data from excel file to create a dataframe"""
        self.df = data_frame
        self._format_measurements()

    def read_excel(self) -> None:
        """Takes data from excel file to create a DataFrame"""
        self.df = pd.read_excel(self.file_path, skiprows=8).astype(str)
        self._format_measurements()

    def read_excel_in_read_mode(self) -> None:
        """Reads excel file in read mode to allow data extraction while excel file is open"""
        # Load workbook
        wb = openpyxl.load_workbook(filename=self.file_path, read_only=True)
        # Select active worksheet
        ws = wb.active
        # Get data from worksheet
        data = ws.values
        # Skip the first 8 rows
        for _ in range(8):
            next(data)
        # Get column names from the next line of data
        cols = next(data)
        # Remove the last 5 None values from cols
        cols = [col for col in cols if col is not None]
        # Get the rest of the data
        data_rows = list(data)
        # Only take the first n columns for each row, where n is the number of columns in cols
        data_rows = [row[:len(cols)] for row in data_rows]
        # Remove rows that are entirely empty
        data_rows = [row for row in data_rows if any(cell is not None for cell in row)]
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=cols)
        # Convert all data types to strings
        self.df = df.astype(str)
        self._format_measurements()

    def format(self) -> None:
        """Changes DataFrame into a standard or template DataFrame that is usable in other classes"""
        formatted_df = pd.DataFrame()
        for column in self.df:
            formatted_df[self.rename_header(column)] = self.df[column]
        self.df = formatted_df

    def create_output(self, file_path: str) -> None:
        """Combines template DataFrame with make manipulated DataFrame"""
        self.df.to_excel('model/temp/temp.xlsx', sheet_name='Sheet1', index=False)

        # Load the source workbook
        source_wb = openpyxl.load_workbook('model/temp/temp.xlsx')
        source_sheet = source_wb.active

        # Load the destination workbook
        destination_wb = openpyxl.load_workbook(self.template_path)
        destination_sheet = destination_wb.active

        # Copy cells from source to destination
        for row in source_sheet.iter_rows(min_row=2, min_col=1):
            for cell in row:
                destination_sheet.cell(row=cell.row + 8, column=cell.column, value=cell.value)

        # Copy cells T1 and U1 from source to T9 and U9 in destination (Grounded and Molded headers)
        destination_sheet['T9'] = source_sheet['T1'].value
        destination_sheet['U9'] = source_sheet['U1'].value

        # Save the destination workbook
        destination_wb.save(file_path)

    def _format_measurements(self):
        """Formats attachment heights to not include decimal"""
        self.df['Neutral'] = pd.to_numeric(self.df['Neutral'], errors='coerce').astype('Int64')
        self.df['Secondary'] = pd.to_numeric(self.df['Secondary'], errors='coerce').astype('Int64')
        self.df['Drip Loop'] = pd.to_numeric(self.df['Drip Loop'], errors='coerce').astype('Int64')
        self.df['Secondary Riser'] = pd.to_numeric(self.df['Secondary Riser'], errors='coerce').astype('Int64')
        self.df['Street Light'] = pd.to_numeric(self.df['Street Light'], errors='coerce').astype('Int64')
        self.df['CATV'] = pd.to_numeric(self.df['CATV'], errors='coerce').astype('Int64')
        self.df['TelCo'] = pd.to_numeric(self.df['TelCo'], errors='coerce').astype('Int64')
        self.df['Fiber'] = pd.to_numeric(self.df['Fiber'], errors='coerce').astype('Int64')
        self.df['Requested Attachment'] = pd.to_numeric(self.df['Requested Attachment'],
                                                        errors='coerce').astype('Int64')
